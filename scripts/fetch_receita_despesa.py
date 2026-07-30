"""
Busca o demonstrativo mais recente de "Estatisticas Fiscais do Governo Geral"
(Tesouro Nacional / STN+IBGE+BCB, metodologia GFSM 2014) e extrai Receita
(abas 3.2/3.2.a anual, 1.3/1.3.a trimestral) e Despesa (abas 3.3/3.3.a anual,
1.4/1.4.a trimestral) do Governo Geral.

Gera dois recortes por bloco (receita/despesa):
  - "anual": serie anual completa, direto das abas 3.2/3.3.
  - "trimestral": serie trimestral crua (sem acumular/suavizar), direto das
    abas 1.3/1.4, indo ate o trimestre corrente disponivel (ex: 2026-I).

Esse dado nao tem API/JSON oficial: e publicado como planilha .xlsx trimestral
no Tesouro Transparente. A pagina de publicacao abaixo e um link "fixo" que
sempre mostra a edicao mais recente (o conteudo eh trocado in-place a cada
trimestre), entao o robo so precisa raspar essa pagina para achar o link de
download vigente -- sem precisar adivinhar datas ou IDs.

Se a raspagem ou o parsing falharem, mantem os dados anteriores (cache), para
a pagina nunca ficar sem dado -- so potencialmente desatualizada.

Rodado pelo GitHub Actions (.github/workflows/update-data.yml).
"""

import json
import re
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import openpyxl
import requests

PUBLICATION_PAGE = (
    "https://www.tesourotransparente.gov.br/publicacoes/"
    "estatisticas-fiscais-do-governo-geral/2021/22"
)
# link do arquivo principal (nao dos anexos "-anexo/"); muda de conteudo a
# cada trimestre mas o padrao da URL de download e sempre este:
DOWNLOAD_RE = re.compile(r"https://thot-arquivos\.tesouro\.gov\.br/publicacao/\d+")

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = REPO_ROOT / "data" / "receita_despesa_data.json"

QUARTER_RE = re.compile(r"^\d{4}-(I|II|III|IV)$")
ROMAN_TO_INT = {"I": 1, "II": 2, "III": 3, "IV": 4}

# codigos de linha (coluna A das abas) que compoem a composicao de cada bloco.
# receita: 1 = total, 11-14 = categorias que somam o total
RECEITA_TOTAL_CODE = "1"
RECEITA_CATEGORIAS = {
    "11": "Impostos",
    "12": "Contribuições sociais",
    "13": "Transferências / Doações",
    "14": "Outras receitas",
}
# despesa: 2M = total (Gasto + Investimento liquido), 21-28 somam "Gasto",
# 31 e o investimento liquido que fecha o total 2M
DESPESA_TOTAL_CODE = "2M"
DESPESA_CATEGORIAS = {
    "21": "Remuneração de empregados",
    "22": "Uso de bens e serviços",
    "23": "Consumo de capital fixo",
    "24": "Juros",
    "25": "Subsídios",
    "26": "Transferências / Doações",
    "27": "Benefícios previdenciários e assistenciais",
    "28": "Outros gastos",
    "31": "Investimento líquido",
}

# abas anuais (inalterado da versao anterior)
SHEETS_ANUAL = {
    "receita": ("3.2", "3.2.a"),
    "despesa": ("3.3", "3.3.a"),
}
# abas trimestrais (serie crua, sem acumular)
SHEETS_TRIMESTRAL = {
    "receita": ("1.3", "1.3.a"),
    "despesa": ("1.4", "1.4.a"),
}


def find_download_url(tries=3, wait=8):
    """Raspa a pagina de publicacao e devolve a URL do xlsx principal vigente."""
    last_err = None
    for attempt in range(1, tries + 1):
        try:
            resp = requests.get(PUBLICATION_PAGE, timeout=25, headers={
                "User-Agent": "Mozilla/5.0 (CIEM dashboard fetch bot)"
            })
            resp.raise_for_status()
            match = DOWNLOAD_RE.search(resp.text)
            if not match:
                raise RuntimeError("link de download nao encontrado na pagina")
            return match.group(0)
        except Exception as e:  # noqa: BLE001
            last_err = e
            print(f"[pagina publicacao] tentativa {attempt}/{tries} falhou: {e}")
            if attempt < tries:
                time.sleep(wait)
    raise RuntimeError(f"falha definitiva ao achar link de download: {last_err}")


def download_workbook(url, tries=3, wait=8):
    last_err = None
    for attempt in range(1, tries + 1):
        try:
            resp = requests.get(url, timeout=60, headers={
                "User-Agent": "Mozilla/5.0 (CIEM dashboard fetch bot)"
            })
            resp.raise_for_status()
            return openpyxl.load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            last_err = e
            print(f"[download xlsx] tentativa {attempt}/{tries} falhou: {e}")
            if attempt < tries:
                time.sleep(wait)
    raise RuntimeError(f"falha definitiva ao baixar xlsx: {last_err}")


def _is_year_header(v):
    return isinstance(v, (int, float)) and 1990 <= v <= 2100


def parse_annual_sheet(ws, total_code, categorias):
    """Le uma aba anual (3.2, 3.2.a, 3.3 ou 3.3.a) e devolve {years:[...], total:[...], series:{codigo:[...]}}."""
    years = None
    total_values = None
    cat_values = {code: None for code in categorias}

    for row in ws.iter_rows(values_only=True):
        code_raw, label = row[0], row[1]
        if code_raw is None:
            continue
        code = str(code_raw).strip()

        if years is None and _is_year_header(row[2]):
            years = [int(y) for y in row[2:] if _is_year_header(y)]
            continue

        if code == total_code:
            total_values = list(row[2:2 + len(years)]) if years else None
        elif code in cat_values:
            cat_values[code] = list(row[2:2 + len(years)]) if years else None

    if years is None or total_values is None:
        raise RuntimeError(f"nao consegui localizar cabecalho/total na aba {ws.title}")
    missing = [c for c, v in cat_values.items() if v is None]
    if missing:
        raise RuntimeError(f"categorias nao encontradas na aba {ws.title}: {missing}")

    return {"years": years, "total": total_values, "series": cat_values}


def parse_quarterly_sheet(ws, total_code, categorias):
    """Le uma aba trimestral (1.3, 1.3.a, 1.4 ou 1.4.a) e devolve
    {quarters:[...], total:[...], series:{codigo:[...]}}, onde quarters
    e uma lista de strings tipo '2010-I'."""
    quarters = None
    total_values = None
    cat_values = {code: None for code in categorias}

    for row in ws.iter_rows(values_only=True):
        code_raw, label = row[0], row[1]
        if code_raw is None:
            continue
        code = str(code_raw).strip()

        if quarters is None and isinstance(row[2], str) and QUARTER_RE.match(row[2].strip()):
            quarters = [str(q).strip() for q in row[2:] if isinstance(q, str) and QUARTER_RE.match(q.strip())]
            continue

        if code == total_code:
            total_values = list(row[2:2 + len(quarters)]) if quarters else None
        elif code in cat_values:
            cat_values[code] = list(row[2:2 + len(quarters)]) if quarters else None

    if quarters is None or total_values is None:
        raise RuntimeError(f"nao consegui localizar cabecalho/total na aba {ws.title}")
    missing = [c for c, v in cat_values.items() if v is None]
    if missing:
        raise RuntimeError(f"categorias nao encontradas na aba {ws.title}: {missing}")

    return {"quarters": quarters, "total": total_values, "series": cat_values}


def build_annual_block(wb, sheet_brl, sheet_pct, total_code, categorias):
    brl = parse_annual_sheet(wb[sheet_brl], total_code, categorias)
    pct = parse_annual_sheet(wb[sheet_pct], total_code, categorias)
    if brl["years"] != pct["years"]:
        raise RuntimeError(f"anos divergentes entre {sheet_brl} e {sheet_pct}")

    categorias_out = [
        {
            "codigo": codigo,
            "nome": nome,
            "brl": brl["series"][codigo],
            "pct": pct["series"][codigo],
        }
        for codigo, nome in categorias.items()
    ]
    return {
        "years": brl["years"],
        "total": {"brl": brl["total"], "pct": pct["total"]},
        "categorias": categorias_out,
    }


def build_quarterly_block(wb, sheet_brl, sheet_pct, total_code, categorias):
    """Le as abas trimestrais cruas (sem acumular/suavizar) e devolve o
    mesmo formato do bloco anual, so que com 'quarters' no lugar de 'years'."""
    brl = parse_quarterly_sheet(wb[sheet_brl], total_code, categorias)
    pct = parse_quarterly_sheet(wb[sheet_pct], total_code, categorias)
    if brl["quarters"] != pct["quarters"]:
        raise RuntimeError(f"trimestres divergentes entre {sheet_brl} e {sheet_pct}")

    categorias_out = [
        {
            "codigo": codigo,
            "nome": nome,
            "brl": brl["series"][codigo],
            "pct": pct["series"][codigo],
        }
        for codigo, nome in categorias.items()
    ]
    return {
        "quarters": brl["quarters"],
        "total": {"brl": brl["total"], "pct": pct["total"]},
        "categorias": categorias_out,
    }


def build_full_block(wb, chave):
    total_code = RECEITA_TOTAL_CODE if chave == "receita" else DESPESA_TOTAL_CODE
    categorias = RECEITA_CATEGORIAS if chave == "receita" else DESPESA_CATEGORIAS

    sheet_brl_a, sheet_pct_a = SHEETS_ANUAL[chave]
    sheet_brl_t, sheet_pct_t = SHEETS_TRIMESTRAL[chave]

    return {
        "anual": build_annual_block(wb, sheet_brl_a, sheet_pct_a, total_code, categorias),
        "trimestral": build_quarterly_block(wb, sheet_brl_t, sheet_pct_t, total_code, categorias),
    }


def load_previous():
    if OUTPUT_PATH.exists():
        try:
            with open(OUTPUT_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:  # noqa: BLE001
            print(f"aviso: nao consegui ler o JSON anterior ({e})")
    return None


def main():
    previous = load_previous()
    any_failure = False
    failed_keys = []
    result = {}
    source_url = None

    try:
        source_url = find_download_url()
        print(f"[link] usando {source_url}")
        wb = download_workbook(source_url)

        result["receita"] = build_full_block(wb, "receita")
        print(f"[receita] OK - anual: {len(result['receita']['anual']['years'])} anos, "
              f"trimestral: {len(result['receita']['trimestral']['quarters'])} trimestres")

        result["despesa"] = build_full_block(wb, "despesa")
        print(f"[despesa] OK - anual: {len(result['despesa']['anual']['years'])} anos, "
              f"trimestral: {len(result['despesa']['trimestral']['quarters'])} trimestres")

    except Exception as e:  # noqa: BLE001
        print(f"[receita_despesa] ERRO DEFINITIVO: {e}")
        any_failure = True
        failed_keys.append("receita_despesa")
        if previous and previous.get("receita") and previous.get("despesa"):
            result["receita"] = previous["receita"]
            result["despesa"] = previous["despesa"]
            source_url = previous.get("source_url")
            print("[receita_despesa] usando cache anterior")
        else:
            result["receita"] = None
            result["despesa"] = None
            print("[receita_despesa] sem cache anterior disponivel, salvando vazio")

    payload = {
        "receita": result["receita"],
        "despesa": result["despesa"],
        "source_url": source_url,
        "source_page": PUBLICATION_PAGE,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "had_failure": any_failure,
        "failed_keys": failed_keys,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"\nSalvo em {OUTPUT_PATH}")
    if any_failure:
        print(f"Atencao: falhas (usando cache): {failed_keys}")


if __name__ == "__main__":
    main()
